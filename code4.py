import pandas as pd
import numpy as np
import re
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side

# Load the marker references from the .xlsx file
marker_df = pd.read_excel(r'C:\Users\chhat\OneDrive\Documents\Project\SOC\Midsem\test\markers.xlsx', engine='openpyxl')

# Clean column names
marker_df.columns = marker_df.columns.str.strip()

# Print column names to ensure they are correct
print("Columns in markers.xlsx:", marker_df.columns.tolist())

# Initialize a dictionary to store timestamps for each marker
marker_timestamps = {marker.strip(): [] for marker in marker_df['Marker']}

# List of text files to read timestamps from
txt_files = [
    r'C:\Users\chhat\OneDrive\Documents\Project\SOC\Midsem\test\soc1-1.txt',
    r'C:\Users\chhat\OneDrive\Documents\Project\SOC\Midsem\test\soc1-2.txt',
    r'C:\Users\chhat\OneDrive\Documents\Project\SOC\Midsem\test\soc1-3.txt',
    r'C:\Users\chhat\OneDrive\Documents\Project\SOC\Midsem\test\soc1-4.txt',
    r'C:\Users\chhat\OneDrive\Documents\Project\SOC\Midsem\test\soc1-5.txt',
    r'C:\Users\chhat\OneDrive\Documents\Project\SOC\Midsem\test\soc1-6.txt'
]

# Read timestamps and match with markers
for txt_file in txt_files:
    if os.path.exists(txt_file):
        with open(txt_file, 'r', encoding='utf-8') as file:
            for line in file:
                match = re.match(r"([\d]+)\[us\]:\s(.+)", line.strip())
                if match:
                    try:
                        timestamp = float(match.group(1))
                        marker = match.group(2).strip()
                        if marker in marker_timestamps:
                            marker_timestamps[marker].append(timestamp)
                            print(f"Added timestamp {timestamp} for marker '{marker}'")
                        else:
                            print(f"⚠️ Marker '{marker}' not found in Excel file!")
                    except ValueError:
                        print(f"⚠️ Invalid timestamp in line: {line.strip()}")
    else:
        print(f"⚠️ File not found: {txt_file}")

# Compute average timestamps
average_timestamps = {
    marker: round(np.mean(timestamps) / 1000) if timestamps else np.nan
    for marker, timestamps in marker_timestamps.items()
}

# Create result DataFrame with threshold and pass/fail status
results = []
for marker in marker_df['Marker']:
    avg_timestamp = average_timestamps.get(marker, np.nan)
    threshold = marker_df.loc[marker_df['Marker'] == marker, 'Threshold'].values[0]
    if not np.isnan(avg_timestamp):
        status = 'Pass' if abs(avg_timestamp - threshold) <= 10 else 'Fail'
    else:
        status = 'No Data'
    results.append({
        'Marker': marker,
        'Threshold': threshold,
        'Average Timestamp(ms)': avg_timestamp,
        'Status': status
    })

# Reorder columns as requested
average_df = pd.DataFrame(results)[['Marker', 'Threshold', 'Average Timestamp(ms)', 'Status']]

# Save the results
output_dir = r'C:\Users\chhat\OneDrive\Documents\Project\SOC\Midsem\test\Result'
os.makedirs(output_dir, exist_ok=True)
output_file = os.path.join(output_dir, "average_timestamps-final.xlsx")
average_df.to_excel(output_file, index=False)

# 🟩 Apply color coding in Excel and add borders
wb = load_workbook(output_file)
ws = wb.active

# Define color fill based on status
status_fill = {
    'Pass': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),  # Green
    'Fail': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),  # Red
    'No Data': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow
}

# Apply colors based on status
for row in range(2, ws.max_row + 1):
    status = ws[f'D{row}'].value
    if status in status_fill:
        ws[f'D{row}'].fill = status_fill[status]

# Define border style (All borders)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Apply borders to the entire table
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.border = thin_border

# Save workbook with updates
wb.save(output_file)
print(f"✅ Excel file saved with color-coded statuses and borders at: {output_file}")